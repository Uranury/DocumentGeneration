import tempfile
import shutil
import json
import re
import logging
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange

logger = logging.getLogger(__name__)

def xlsx_copy_cell_style(src_cell, dst_cell):
    """Copy cell styling from source to destination cell"""
    if src_cell.font:
        dst_cell.font = src_cell.font.copy()
    if src_cell.border:
        dst_cell.border = src_cell.border.copy()
    if src_cell.fill:
        dst_cell.fill = src_cell.fill.copy()
    if src_cell.alignment:
        dst_cell.alignment = src_cell.alignment.copy()
    if src_cell.number_format:
        dst_cell.number_format = src_cell.number_format

def xlsx_get_placeholders(ws, table_name):
    """Find all placeholders for a specific table in the worksheet"""
    placeholder_cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                # Match patterns like {{ table_name.field_name }}
                match = re.match(r"\{\{\s*([a-zA-Z0-9_]+)\.([a-zA-Z0-9_]+)\s*\}\}", cell.value)
                if match:
                    t_name, field_name = match.groups()
                    if t_name == table_name:
                        if table_name not in placeholder_cells:
                            placeholder_cells[table_name] = {}
                        placeholder_cells[table_name][field_name] = (cell.row, cell.column)
    return placeholder_cells

def insert_rows_and_shift_merges(ws, start_row, num_rows):
    """Insert rows and properly handle merged cells"""
    if num_rows <= 0:
        return
        
    old_merges = list(ws.merged_cells.ranges)
    
    # Insert rows
    ws.insert_rows(start_row, amount=num_rows)
    
    # Create all cells in merges manually to avoid KeyError
    for rng in old_merges:
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                if (row, col) not in ws._cells:
                    ws._cells[(row, col)] = ws.cell(row=row, column=col)
    
    # Safely remove old merges
    for rng in old_merges:
        try:
            ws.unmerge_cells(str(rng))
        except KeyError:
            # Ignore uninitialized cells
            pass
    
    # Re-add shifted merges
    for rng in old_merges:
        if rng.min_row >= start_row:
            shifted = CellRange(
                min_col=rng.min_col,
                min_row=rng.min_row + num_rows,
                max_col=rng.max_col,
                max_row=rng.max_row + num_rows,
            )
            ws.merge_cells(str(shifted))
        else:
            ws.merge_cells(str(rng))

def replace_placeholders(text, data_dict):
    """Replace simple {{ variable }} placeholders in text"""
    if not isinstance(text, str):
        return text
    
    # Pattern for simple variables like {{ variable_name }}
    pattern = r'\{\{\s*([a-zA-Z0-9_]+)\s*\}\}'
    
    def replace_var(match):
        var_name = match.group(1).strip()
        if var_name in data_dict and not isinstance(data_dict[var_name], list):
            return str(data_dict[var_name])
        return match.group(0)  # Return original if not found or is a list
    
    return re.sub(pattern, replace_var, text)

def process_xlsx(ws, data_dict):
    """Process XLSX worksheet with data substitution and table handling"""
    try:
        # Validate input data first
        if not data_dict:
            raise ValueError("Empty data payload")
        
        # Process simple placeholders first (non-table data)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    cell.value = replace_placeholders(cell.value, data_dict)
        
        # Process table data (lists)
        for table_name, table_data in data_dict.items():
            if isinstance(table_data, list) and table_data:
                placeholder_cells = xlsx_get_placeholders(ws, table_name)
                
                if table_name in placeholder_cells:
                    # Find the starting row for this table
                    start_row = min(row for row, col in placeholder_cells[table_name].values())
                    num_rows = len(table_data)
                    
                    # Insert additional rows if we need more than 1
                    if num_rows > 1:
                        insert_rows_and_shift_merges(ws, start_row + 1, num_rows - 1)
                    
                    # Fill in the data
                    for row_idx, row_data in enumerate(table_data, start=start_row):
                        for field_name, (orig_row, col) in placeholder_cells[table_name].items():
                            if field_name in row_data:
                                # Get the original cell for style copying
                                cell_orig = ws.cell(row=start_row, column=col)
                                # Create/update the target cell
                                cell_new = ws.cell(row=row_idx, column=col, value=row_data[field_name])
                                # Copy styling from original cell
                                xlsx_copy_cell_style(cell_orig, cell_new)
                                
    except Exception as e:
        logger.error(f"Error processing XLSX template: {e}")
        raise Exception(f"Error processing spreadsheet data: {e}")

def render_xlsx_file(template_file, data_dict):
    """
    Renders an XLSX template with data using openpyxl with advanced table support.
    
    Template syntax examples:
    - Simple variables: {{ variable_name }}
    - Table data: {{ table_name.field_name }}
    
    For tables, all {{ table_name.field_name }} placeholders in the same row
    will be replicated for each item in the table_name list.
    """
    import tempfile
    import shutil
    
    # Save uploaded file to temporary location
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_template:
        shutil.copyfileobj(template_file, tmp_template)
        tmp_template_path = tmp_template.name
    
    # Create output file
    output_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    output_path = output_file.name
    output_file.close()
    
    # Process the template
    wb = load_workbook(tmp_template_path)
    try:
        ws = wb.active
        process_xlsx(ws, data_dict)
        wb.save(output_path)
    finally:
        wb.close()
    
    return output_path

